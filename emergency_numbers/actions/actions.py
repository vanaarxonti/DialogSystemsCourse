
import os
import pandas as pd
from typing import Any, Dict, List, Text
from rasa_sdk import Action, Tracker
from rasa_sdk.executor import CollectingDispatcher
from rasa_sdk.events import AllSlotsReset
import requests
from bs4 import BeautifulSoup
from rasa_sdk import Action
from rasa_sdk.events import SlotSet
from rasa_sdk import Action
from rasa_sdk.events import SlotSet



class ActionResetSlots(Action):

    def name(self) -> Text:
         return "action_reset_slots"

    def run(self, dispatcher: CollectingDispatcher,
             tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

           dispatcher.utter_message(text="I am waiting for new info")

           return [AllSlotsReset()]
    


from openpyxl import load_workbook

class ActionEmergencyNumbers(Action):

    def name(self) -> Text:
        return "action_emergency_numbers"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        

        country = tracker.get_slot("country")
        wb = load_workbook("ARES.xlsx")
        sheet = wb["Sheet1"]
        from fuzzywuzzy import fuzz

        similarity_threshold = 85

        for row in range(2, sheet.max_row + 1):
            similarity = fuzz.token_set_ratio(sheet.cell(row, 2).value, country)
            if similarity > similarity_threshold:
                police = sheet.cell(row, 3).value
                ambulance = sheet.cell(row, 4).value
                fire = sheet.cell(row, 5).value
                break

        else:
            dispatcher.utter_message(f"Sorry, I couldn't find emergency numbers for {country}")
            return []
        
        message = f"Emergency numbers for {country} are the following:\nPolice: {police}\nAmbulance: {ambulance}\nFire: {fire}. "
        dispatcher.utter_message(message)

